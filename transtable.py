# 対訳ファイル作成アプリ
# 和文と英文の統合報告書PDFから対訳ファイルのエクセルを作成します
# 2026-02-26
# openpyxlライブラリのインストールが必要
# 使い方
# 1. 和文PDFをアップロード
# 2. 英文PDFをアップロード
# 3. 対訳ファイルが作成されるので、ダウンロードボタンを押して保存してください

import fitz
import pandas as pd
import streamlit as st
import os
import re
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from io import BytesIO

def split_spread_pdf(input_pdf):
    """見開きPDFをメモリ上で分割処理し、分割済みPDFを返す"""
    # BytesIO または UploadedFile オブジェクトを処理
    if hasattr(input_pdf, 'read'):
        input_bytes = BytesIO(input_pdf.read())
    else:
        with open(input_pdf, 'rb') as f:
            input_bytes = BytesIO(f.read())
    
    doc = fitz.open(stream=input_bytes.getvalue(), filetype="pdf")
    new_doc = fitz.open()

    for page in doc:
        rect = page.rect
        width = rect.width
        height = rect.height

        # 左半分
        left_page = new_doc.new_page(width=width/2, height=height)
        left_page.show_pdf_page(left_page.rect, doc, page.number, 
                                 clip=fitz.Rect(0, 0, width/2, height))

        # 右半分
        right_page = new_doc.new_page(width=width/2, height=height)
        right_page.show_pdf_page(right_page.rect, doc, page.number, 
                                  clip=fitz.Rect(width/2, 0, width, height))

    doc.close()
    
    # メモリ上のBytesIOに保存
    output_bytes = BytesIO()
    new_doc.save(output_bytes)
    new_doc.close()
    
    output_bytes.seek(0)
    return output_bytes


def merge_paragraphs(blocks, threshold=5):
    """
    隣接するブロックを結合して段落を復元する
    threshold: Y座標の差がこの値以下のブロックは同じ段落として結合する
    """
    if not blocks:
        return []
    
    merged = []
    current_paragraph = {
        'text': blocks[0][4].strip(),
        'y0': blocks[0][1],
        'y1': blocks[0][3]
    }
    
    for i in range(1, len(blocks)):
        block = blocks[i]
        text = block[4].strip()
        y0 = block[1]
        y1 = block[3]
        
        # 前のブロックからの距離をチェック
        # 1. Y座標が近い（同じ行付近）
        # 2. 前のテキストが句点で終わっていない
        # 3. 数字で始まっていない
        prev_text = current_paragraph['text']
        
        if (y0 - current_paragraph['y1'] < threshold and 
            text and 
            not any(prev_text.endswith(mark) for mark in ['。', '！', '？']) and
            not re.match(r'^[0-9０-９]', text)):
            # 同じ段落として結合
            current_paragraph['text'] += text
            current_paragraph['y1'] = y1
        else:
            # 新しい段落として記録
            if current_paragraph['text']:
                merged.append(current_paragraph['text'])
            current_paragraph = {
                'text': text,
                'y0': y0,
                'y1': y1
            }
    
    # 最後のパラグラフを追加
    if current_paragraph['text']:
        merged.append(current_paragraph['text'])
    
    return merged


def extract_paragraphs_to_file(doc_ja, doc_en, output_xlsx):
		
    df = pd.DataFrame()

	# 総ページ数を取得
    total_pages = doc_ja.page_count
    # ページごとにループ
    for page_index in range(total_pages):
        list_ja, list_en = [], []
        page_ja = doc_ja[page_index]
        #blocks_ja = page_ja.get_text("blocks", sort=True)
        blocks_ja = page_ja.get_text("blocks")
        merged_ja = merge_paragraphs(blocks_ja)
        list_ja.append("P"+str(page_index)+"\n"+"\n"+"\n")
        #pattern_count = len(re.findall(r'\b\w+,\s*\d+', paras_ja))
        #if pattern_count > threshold:
        #    continue
        for para_ja in merged_ja:
            para_ja = para_ja.strip()
            if para_ja and not re.match(r'^\d+\w*[\s\W]+\d+$', para_ja):
                list_ja.append("\n")
                list_ja.append(para_ja.strip()) 

        page_en = doc_en[page_index]
        #blocks_en = page_en.get_text("blocks", sort=True)
        blocks_en = page_en.get_text("blocks")
        list_en.append("P"+str(page_index)+"\n"+"\n"+"\n")
        #pattern_count = len(re.findall(r'\b\w+,\s*\d+', paras_en))
        #if pattern_count > threshold:
        #    continue
        for block_en in blocks_en:
            para_en = block_en[4].strip()
            if para_en and not re.match(r'^\d+\w*[\s\W]+\d+$', para_en):
                list_en.append("\n")
                list_en.append(para_en.strip()) 

        # 各リストをそれぞれ一意のカラム名でDataFrame化
        df_ja = pd.DataFrame({'ja': list_ja})
        df_en = pd.DataFrame({'en': list_en})
        # インデックスで揃えて結合（列名がユニークなので再インデックスエラーを回避）
        df_m = pd.concat([df_ja, df_en], axis=1)
        # 全ページ分を連結する際はインデックスを振り直す
        df = pd.concat([df, df_m], ignore_index=True)
        
    #df.to_excel(output_xlsx, index=False)
    df.to_excel(output_xlsx, engine='xlsxwriter', index=False)
    
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb.active
    # A列とB列の幅を設定
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 100

    # 全セルに折り返し（wrap_text）とフォントサイズを設定
    default_font_size = 14
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrapText=True, vertical='top')
            cell.font = Font(size=default_font_size)

    # 行高さを改良して自動調整（改行・単語ラップ・フォントサイズを考慮）
    line_height = default_font_size * 1.2  # おおよその1行あたりの高さ（pt）
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            text = str(cell.value)
            paragraphs = text.split('\n')
            col_letter = get_column_letter(col_idx)
            col_width = ws.column_dimensions[col_letter].width or 10
            # 列幅から1行に入ると推定される文字数を算出（経験則として係数を掛ける）
            est_chars_per_line = max(10, int(col_width * 1.8))

            # 単語単位でラップをシミュレートして必要行数を数える
            lines = 0
            for para in paragraphs:
                words = para.split()
                if not words:
                    lines += 1
                    continue
                cur_len = 0
                for w in words:
                    wl = len(w)
                    if cur_len == 0:
                        cur_len = wl
                    elif cur_len + 1 + wl <= est_chars_per_line:
                        cur_len += 1 + wl
                    else:
                        lines += 1
                        cur_len = wl
                if cur_len > 0:
                    lines += 1

            if lines > max_lines:
                max_lines = lines

        # 行高さを設定（最小15pt、最大はコンテンツに応じたサイズ）
        ws.row_dimensions[row_idx].height = max(15, max_lines * line_height)
    
    # 交互の色の定義
    #fill_even = PatternFill(patternType='solid', fgColor='CEE6C1') # 薄い緑
    #fill_odd = PatternFill(patternType='solid', fgColor='FFFFFF')  # 白

    # 各行に色を適用
    #for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
    #    if row[0].row % 2 == 0:
    #        for cell in row:
    #            cell.fill = fill_even
    #    else:
    #        for cell in row:
    #            cell.fill = fill_odd

    wb.save(output_xlsx)

    print("処理が完了しました")


st.title("対訳ファイル作成")
st.write("和文と英文の統合報告書PDFから対訳ファイルのエクセルを作成します")

# セッション状態の初期化
if 'processing_done' not in st.session_state:
    st.session_state.processing_done = False
if 'output_xlsx' not in st.session_state:
    st.session_state.output_xlsx = None

spread = st.checkbox("PDFが見開きの場合、チェックしてください", value=True)

uploaded_file_ja = st.file_uploader("和文（PDFファイル）をアップロード", type="pdf", key="file_ja")
doc_ja = None
if uploaded_file_ja is not None:
    file_name = os.path.splitext(uploaded_file_ja.name)[0]  # アップロードされたファイル名を取得
    if spread:
        pdf_bytes_ja = split_spread_pdf(uploaded_file_ja)
        doc_ja = fitz.open(stream=pdf_bytes_ja.getvalue(), filetype="pdf")
    else:
        doc_ja = fitz.open(stream=uploaded_file_ja.read(), filetype="pdf")
    st.success("和文ファイルがアップロードされました。")

uploaded_file_en = st.file_uploader("英文（PDFファイル）をアップロード", type="pdf", key="file_en")
doc_en = None
if uploaded_file_en is not None:
    if spread:
        pdf_bytes_en = split_spread_pdf(uploaded_file_en)
        doc_en = fitz.open(stream=pdf_bytes_en.getvalue(), filetype="pdf")
    else:
        doc_en = fitz.open(stream=uploaded_file_en.read(), filetype="pdf")
    st.success("英文ファイルがアップロードされました。")
    
    # 初回実行時のみ処理を実行
    if not st.session_state.processing_done:
        try:
            output_xlsx = f"{file_name}_output.xlsx"  # アップロードされたファイル名をベースに出力ファイル名を作成
            extract_paragraphs_to_file(doc_ja, doc_en, output_xlsx)
            st.success(f"対訳ファイルが作成されました: {output_xlsx}")
            st.session_state.processing_done = True
            st.session_state.output_xlsx = output_xlsx

        except Exception as e:
            st.error(f"ファイルの読み込み中にエラーが発生しました: {e}")
        
        finally:
            # ドキュメントを閉じる
            if doc_ja:
                doc_ja.close()
            if doc_en:
                doc_en.close()
    
    # 処理済みの場合はダウンロードボタンを表示
    if st.session_state.processing_done:
        st.success("ダウンロードボタンを押してください")        
        with open(st.session_state.output_xlsx, "rb") as file:
            xlsx_data = file.read()
            # ダウンロードボタンを作成
            st.download_button(
                label="Excelをダウンロード",
                data=xlsx_data,
                file_name=file_name+"_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
else:
    st.info("ファイルをアップロードしてください。")
